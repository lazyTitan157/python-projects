import requests
import json
from openpyxl import Workbook
import csv

# API Settings
url = "https://api.upbit.com/v1/candles/days"
headers = {"Accept": "application/json"}
to = "2021-05-25T00:00:00Z"

# Excel file Settings
write_wb = Workbook()
write_ws = write_wb.create_sheet('KRW-BTC')
write_ws = write_wb.active

# CSV file Settings
file = open("KRW_BTC.csv", mode="w", encoding="utf-8", newline="")
writer = csv.writer(file)

# csv data init
dayline = ["Date"]
dataline = ["Data"]

# 요청 횟수 제한 주의하기 (https://docs.upbit.com/docs/user-request-guide#quotation-api)
# 최신순으로 데이터 저장
for i in range(1, 6):
    querystring = {"market":"KRW-BTC","to":to, "count":"200"} # count max = 200
    response = requests.request("GET", url, headers=headers, params=querystring)

    if response.status_code == requests.codes.ok:
        print("API Response Status code: ok")
        
        response_native = json.loads(response.text)

        
        for index, item in enumerate(response_native):
            if index == len(response_native)-1:
                to = item.get('candle_date_time_kst')+'Z'
                print('Next: ',to)

            # data add
            write_ws.cell(1, 200*(i-1)+index+1, item.get('candle_date_time_kst'))
            write_ws.cell(2, 200*(i-1)+index+1, item.get('trade_price'))

            # csv data add
            dayline.append(item.get('candle_date_time_kst'))
            dataline.append(item.get('trade_price'))

# csv file save
writer.writerow(dayline)
writer.writerow(dataline)

# csv save end
file.close()

# Excel file save
write_wb.save("C:/projects/toyproject/kaist/KRW_BTC.xlsx")

print("end")
